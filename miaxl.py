import datetime
import openpyxl

# ----------------------- This is 79 characters -------------------------------

wb = openpyxl.load_workbook('feb2017.xlsx')
ws = wb.worksheets[0]
maxc = ws.max_column
maxr = ws.max_row

for row in ws.iter_rows(min_row=1, max_col=maxc, max_row=maxr):
    for cell in row:
        print(cell.value, end=' ')
    print()

