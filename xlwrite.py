import openpyxl

wb = openpyxl.load_workbook('six_flags_registration.xlsx')
sheet1 = wb.create_sheet("Sheet1")
sheet2 = wb.create_sheet("Sheet2")
sheet3 = wb.create_sheet("Sheet3")



